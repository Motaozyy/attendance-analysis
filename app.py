"""
Streamlit 考勤分析看板 - 基于打卡时间数据
三大分析视角：离职前对比 / 同部门对比 / 个人趋势
三模型集成：RF + XGBoost + Logistic Regression
"""
import streamlit as st
import pandas as pd
import numpy as np
import sqlite3
import plotly.express as px
import plotly.graph_objects as go
import os
import joblib
import tempfile
from datetime import datetime
from train_models import monthly_agg, predict_employee

st.set_page_config(page_title="考勤分析看板", layout="wide")

DB_PATH = os.path.join(os.path.dirname(__file__), 'attendance.db')
MODEL_PATH = os.path.join(os.path.dirname(__file__), 'models', 'attendance_models.pkl')
MODEL_DIR = os.path.join(os.path.dirname(__file__), 'models')

# ── session_state 初始化 ──
for key in ['upload_counter_full', 'upload_counter_incr', 'upload_success_msg']:
    if key not in st.session_state:
        st.session_state[key] = '' if key == 'upload_success_msg' else 0


# =====================================================================
#  带进度回调的上传处理
# =====================================================================
def process_uploaded_file(uploaded_file, progress=None, status=None):
    """处理用户上传的Excel文件，重建数据库和模型"""
    from time_parser import parse_clock_data
    import train_models as tm

    # 获取原始文件名扩展名
    original_name = uploaded_file.name.lower()
    suffix = '.xls'

    if progress:
        progress(10)
    if status:
        status("保存上传文件...")

    # 保存上传的文件到临时路径
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(uploaded_file.getvalue())
        tmp_path = tmp.name

    try:
        # 如果是 .xlsx，先转换
        if original_name.endswith('.xlsx'):
            if progress:
                progress(20)
            if status:
                status("转换 .xlsx → .xls 格式...")
            import openpyxl
            import xlwt
            wb_in = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
            wb_out = xlwt.Workbook(encoding='utf-8')

            for sheet_name in wb_in.sheetnames:
                ws_in = wb_in[sheet_name]
                ws_out = wb_out.add_sheet(sheet_name, cell_overwrite_ok=True)
                for r, row in enumerate(ws_in.iter_rows(values_only=True)):
                    for c, val in enumerate(row):
                        if val is not None:
                            ws_out.write(r, c, val)

            xls_path = tmp_path.replace('.xls', '_converted.xls')
            wb_out.save(xls_path)
            os.unlink(tmp_path)
            tmp_path = xls_path

        if progress:
            progress(40)
        if status:
            status("解析打卡数据 (同步工作日/休息日)...")

        # 解析数据
        parse_clock_data(file_path=tmp_path)

        if progress:
            progress(60)
        if status:
            status("提取特征...")

        # 重新训练模型
        try:
            df_feats = tm.extract_features()
            if len(df_feats) < 5:
                return True, "数据解析成功！但样本量不足（<5人），未重新训练模型，可直接查看打卡数据。"
            n_resigned = df_feats['is_resigned'].sum()
            if n_resigned < 2:
                return True, f"数据解析成功！但离职员工仅{n_resigned:.0f}人，不足以训练模型，可直接查看打卡数据。"

            if progress:
                progress(75)
            if status:
                status(f"训练三模型 (RF+XGB+LR) / {int(len(df_feats))}人样本...")

            tm.train_models(df_feats)
        except Exception as model_err:
            err_str = str(model_err)
            if 'only 1 member' in err_str or 'only one class' in err_str:
                return True, "数据解析成功！但上传数据中无离职员工样本，无法训练预测模型，可直接查看打卡数据。"
            return True, f"数据解析成功！但模型训练失败，可直接查看打卡数据。"

        if progress:
            progress(100)

        return True, "数据处理和模型训练完成！"
    except Exception as e:
        return False, f"处理出错: {str(e)}"
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)


def process_incremental_upload(uploaded_file, progress=None, status=None):
    """增量上传：仅将新月份数据追加到现有数据库"""
    from time_parser import parse_clock_data
    import uuid

    if progress:
        progress(10)
    if status:
        status("保存上传文件...")

    # 保存上传文件到临时路径
    original_name = uploaded_file.name.lower()
    suffix = '.xls'
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(uploaded_file.getvalue())
        tmp_path = tmp.name

    try:
        # .xlsx 转换
        if original_name.endswith('.xlsx'):
            if progress:
                progress(20)
            if status:
                status("转换 .xlsx → .xls 格式...")
            import openpyxl, xlwt
            wb_in = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
            wb_out = xlwt.Workbook(encoding='utf-8')
            for sheet_name in wb_in.sheetnames:
                ws_in = wb_in[sheet_name]
                ws_out = wb_out.add_sheet(sheet_name, cell_overwrite_ok=True)
                for r, row in enumerate(ws_in.iter_rows(values_only=True)):
                    for c, val in enumerate(row):
                        if val is not None:
                            ws_out.write(r, c, val)
            xls_path = tmp_path.replace('.xls', '_converted.xls')
            wb_out.save(xls_path)
            os.unlink(tmp_path)
            tmp_path = xls_path

        if progress:
            progress(35)
        if status:
            status("解析数据 → 临时数据库...")

        # 解析到临时数据库
        tmp_db = os.path.join(tempfile.gettempdir(), f'attendance_incr_{uuid.uuid4().hex}.db')
        parse_clock_data(file_path=tmp_path, db_path=tmp_db)

        if progress:
            progress(55)
        if status:
            status("比对月份差异...")

        # 读取临时数据库的新数据
        conn_tmp = sqlite3.connect(tmp_db)
        new_clock = pd.read_sql("SELECT * FROM clock_records", conn_tmp)
        new_roster = pd.read_sql("SELECT * FROM employee_roster", conn_tmp)
        conn_tmp.close()
        os.unlink(tmp_db)

        if len(new_clock) == 0:
            return False, "上传文件中没有有效的打卡数据。"

        # 读取现有数据库
        conn_main = sqlite3.connect(DB_PATH)
        exist_clock = pd.read_sql("SELECT * FROM clock_records", conn_main)
        exist_roster = pd.read_sql("SELECT * FROM employee_roster", conn_main)

        # 保存原始列名
        orig_cols = list(exist_clock.columns)

        # 统一转换日期
        exist_clock['date'] = pd.to_datetime(exist_clock['date'])
        new_clock['date'] = pd.to_datetime(new_clock['date'])
        exist_clock['ym_key'] = exist_clock['date'].dt.year.astype(int) * 100 + exist_clock['date'].dt.month.astype(int)
        new_clock['ym_key'] = new_clock['date'].dt.year.astype(int) * 100 + new_clock['date'].dt.month.astype(int)

        new_ym_set = set(new_clock['ym_key'].unique())
        exist_ym_set = set(exist_clock['ym_key'].unique())

        # 需要覆盖的月份（上传文件中有的月份，不论是否已存在）
        # 策略：删除已有数据库中这些月份的全部记录，再用上传的替换
        overwrite_ym_set = new_ym_set & exist_ym_set  # 上传和数据库都有的月份 → 覆盖
        fresh_ym_set = new_ym_set - exist_ym_set      # 数据库没有的月份 → 新增

        total_affected = overwrite_ym_set | fresh_ym_set

        if not total_affected:
            conn_main.close()
            return False, "上传文件中没有有效月份数据。"

        # 从已存在数据中剔除"将被覆盖"的月份
        keep_mask = ~exist_clock['ym_key'].isin(overwrite_ym_set)
        exist_clock_keep = exist_clock[keep_mask].copy()

        # 新数据只取上传文件中的月份
        incr_clock = new_clock[new_clock['ym_key'].isin(total_affected)].copy()

        # 去掉临时列
        incr_clock = incr_clock.drop(columns=['ym_key'], errors='ignore')
        exist_clock_keep = exist_clock_keep.drop(columns=['ym_key'], errors='ignore')

        if progress:
            progress(70)
        if status:
            status("合并数据 → 写回数据库...")

        # 合并
        merged_clock = pd.concat([exist_clock_keep, incr_clock], ignore_index=True)
        merged_clock = merged_clock[orig_cols]

        # 合并员工花名册：新增或更新
        exist_names = set(exist_roster['name'])
        new_names = new_roster[~new_roster['name'].isin(exist_names)]
        merged_roster = pd.concat([exist_roster, new_names], ignore_index=True)

        conn_main.execute("DROP TABLE IF EXISTS clock_records")
        conn_main.execute("DROP TABLE IF EXISTS employee_roster")
        merged_clock.to_sql('clock_records', conn_main, if_exists='replace', index=False)
        merged_roster.to_sql('employee_roster', conn_main, if_exists='replace', index=False)

        if progress:
            progress(80)
        if status:
            status("创建索引...")
        conn_main.execute('CREATE INDEX idx_clock_name ON clock_records(name)')
        conn_main.execute('CREATE INDEX idx_clock_date ON clock_records(date)')
        conn_main.execute('CREATE INDEX idx_roster_name ON employee_roster(name)')
        conn_main.commit()
        conn_main.close()

        # 增量上传不重新训练模型，直接使用已有模型
        if progress:
            progress(100)
        if status:
            status("完成！")

        overwrite_ym_str = sorted([f'{k//100}-{k%100:02d}' for k in overwrite_ym_set]) if overwrite_ym_set else []
        fresh_ym_str = sorted([f'{k//100}-{k%100:02d}' for k in fresh_ym_set]) if fresh_ym_set else []

        msg_parts = [f"共处理 {len(incr_clock)} 条打卡记录"]
        if overwrite_ym_str:
            msg_parts.append(f"覆盖月份: {', '.join(overwrite_ym_str)}")
        if fresh_ym_str:
            msg_parts.append(f"新增月份: {', '.join(fresh_ym_str)}")

        return True, "增量上传成功！" + " | ".join(msg_parts)
    except Exception as e:
        return False, f"增量上传处理出错: {str(e)}"
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)


# =====================================================================
#  数据加载
# =====================================================================
@st.cache_data
def load_data():
    """加载数据库数据，若数据库不存在则返回空 DataFrame"""
    if not os.path.exists(DB_PATH):
        return pd.DataFrame(), pd.DataFrame()
    try:
        conn = sqlite3.connect(DB_PATH)
        df_clock = pd.read_sql("SELECT * FROM clock_records ORDER BY date", conn)
        df_roster = pd.read_sql("SELECT * FROM employee_roster", conn)
        conn.close()

        if not df_clock.empty:
            df_clock['date'] = pd.to_datetime(df_clock['date'])
            df_clock['first_time_str'] = df_clock['first_time_str'].fillna('')
            df_clock['last_time_str'] = df_clock['last_time_str'].fillna('')
        return df_clock, df_roster
    except Exception:
        return pd.DataFrame(), pd.DataFrame()


def load_model():
    """从磁盘加载模型。
    不使用缓存保证上传新数据后模型信息立即更新。
    """
    if not os.path.exists(MODEL_PATH):
        return None
    return joblib.load(MODEL_PATH)


df_clock, df_roster = load_data()
model_info = load_model()

# 在职员工列表
if not df_roster.empty and 'status' in df_roster.columns:
    active_roster = df_roster[df_roster['status'] == '在职']
    active_names = set(active_roster['name'].unique())
else:
    active_roster = pd.DataFrame()
    active_names = set()


# =====================================================================
#  侧边栏导航
# =====================================================================
st.sidebar.title("考勤分析系统")

# ── 格式说明 & 模板下载 ──
with st.sidebar.expander("格式说明 & 下载模板", expanded=False):
    st.markdown("""
    **Excel 文件要求：**
    1. 需包含 **两个 Sheet**：
       - `考勤表`：原始打卡时间数据
       - `花名册`：员工信息
    2. 考勤表格式：每月一块，标题行含 `打卡时间 统计日期`，
       日期列标记不限（纯数字1~31、或"六""日""端午"等中文均可），
       系统会调用**中国法定节假日日历**自动识别工作日/休息日。
    3. 员工数据行：工作日填打卡时间，休息日留空

    **花名册字段说明（按列顺序）：**

    | 列号 | 字段名 | 说明 |
    |:---:|:------|:-----|
    | A | 姓名 | 任意文本，须与考勤表一致 |
    | B | 性别 | 仅展示 |
    | C | 部门 | **核心字段**，按部门分组分析 |
    | D | 岗位 | 仅展示 |
    | E | 职级 | 仅展示 |
    | F | 员工状态 | **必须写`在职`或`离职`** |
    | G | 员工类型 | 仅展示 |
    | H | 归属条线 | 仅展示 |
    | I | 进入公司时间 | YYYY-MM-DD 或 Excel日期值 |

    **考勤表字段说明：**

    | 列号 | 字段名 | 格式 |
    |:---:|:------|:-----|
    | A | 姓名 | 与花名册一致 |
    | B | 考勤组 | 任意 |
    | C | 部门 | 任意 |
    | D | 工号 | 任意 |
    | E | 职位 | 任意 |
    | F | UserId | 任意 |
    | G~ | 日期列 | **休息日留空**；**工作日填 `HH:MM\\nHH:MM`** |

    > ⚠️ 打卡时间必须为 24小时制 `HH:MM`，多笔用换行符分隔（`\\n`）。
    > 12:00前取最早时间为上班打卡，12:00后取最晚时间为下班打卡。
    """)
    template_path = os.path.join(os.path.dirname(__file__), '考勤样例模板.xls')
    if os.path.exists(template_path):
        with open(template_path, 'rb') as f:
            st.download_button(
                label="下载样例模板",
                data=f,
                file_name="考勤样例模板.xls",
                mime="application/vnd.ms-excel",
                help="5名员工、3个月数据，系统自动识别工作日/休息日，与原始数据格式一致",
            )

# ── 上传数据：全量替换 ──
with st.sidebar.expander("📦 全量替换数据", expanded=False):
    st.warning("⚠️ 将**完全替换**现有数据库，不会保留原有数据。")
    upload_key_full = f"global_upload_full_{st.session_state['upload_counter_full']}"
    uploaded_full = st.file_uploader(
        "选择完整考勤Excel",
        type=['xls', 'xlsx'],
        key=upload_key_full,
        help="格式与样例模板一致"
    )
    if uploaded_full is not None:
        prog = st.progress(0)
        status_text = st.empty()
        with status_text.container():
            st.caption("保存上传文件...")

        success, msg = process_uploaded_file(
            uploaded_full,
            progress=lambda v: prog.progress(v),
            status=lambda s: status_text.caption(s)
        )
        if success:
            st.session_state['upload_success_msg'] = msg
            st.session_state['upload_counter_full'] += 1
            st.cache_data.clear()
            st.cache_resource.clear()
            st.rerun()
        else:
            prog.empty()
            status_text.empty()
            st.error(msg)

# ── 上传数据：增量追加 ──
with st.sidebar.expander("📈 增量追加月份", expanded=False):
    st.info("上传文件中的月份，**已有则覆盖、没有则新增**。适合补录或修正数据。")
    upload_key_incr = f"global_upload_incr_{st.session_state['upload_counter_incr']}"
    uploaded_incr = st.file_uploader(
        "选择含新月份的Excel",
        type=['xls', 'xlsx'],
        key=upload_key_incr,
        help="上传文件中所有月份均会写入：已有月份覆盖，新月份追加"
    )
    if uploaded_incr is not None:
        prog = st.progress(0)
        status_text = st.empty()
        with status_text.container():
            st.caption("保存上传文件...")

        success, msg = process_incremental_upload(
            uploaded_incr,
            progress=lambda v: prog.progress(v),
            status=lambda s: status_text.caption(s)
        )
        if success:
            st.session_state['upload_success_msg'] = msg
            st.session_state['upload_counter_incr'] += 1
            st.cache_data.clear()
            st.cache_resource.clear()
            st.rerun()
        else:
            prog.empty()
            status_text.empty()
            st.error(msg)

# ── 显示成功通知（跨页面持久） ──
if st.session_state['upload_success_msg']:
    msg = st.session_state['upload_success_msg']
    st.session_state['upload_success_msg'] = ''
    st.toast(msg, icon='✅')

page = st.sidebar.radio("导航", [
    "数据概览",
    "部门考勤看板",
    "员工考勤分析",
    "离职风险预测",
    "模型信息",
])

st.sidebar.markdown("---")
st.sidebar.caption(
    "**数据说明**\n\n"
    "基于打卡时间数据，以12:00为分界：\n"
    "**上班（12:00前）**\n"
    "- 早到：< 08:30\n"
    "- 正常上班：08:30 - 09:00\n"
    "- 迟到：> 09:00\n"
    "- 上班缺卡：无12:00前打卡\n"
    "**下班（12:00后）**\n"
    "- 早退：< 17:30\n"
    "- 正常下班：17:30 - 18:00\n"
    "- 加班：> 18:00\n"
    "- 下班缺卡：无12:00后打卡\n"
    "- 当天无打卡：缺勤"
)


# =====================================================================
#  MORNING/EVENING 颜色方案
# =====================================================================
MORNING_COLORS = {
    '早到': '#3498db',
    '正常上班': '#2ecc71',
    '迟到': '#e74c3c',
    '上班缺卡': '#95a5a6',
}
EVENING_COLORS = {
    '早退': '#e74c3c',
    '正常下班': '#2ecc71',
    '加班': '#9b59b6',
    '下班缺卡': '#95a5a6',
}
# 12:00 分界规则说明
RULE_EXPLANATION = """
**打卡规则（以 12:00 为分界）**
- 12:00 前打卡 → 上班；12:00（含）后打卡 → 下班
- 上班：早到(<08:30) / 正常上班(08:30-09:00) / 迟到(>09:00)
- 下班：早退(<17:30) / 正常下班(17:30-18:00) / 加班(>18:00)
- 仅有上班打卡 → 下班缺卡；仅有下班打卡 → 上班缺卡
- 当天无任何打卡 → 缺勤（不计入记录）
"""


# =====================================================================
#  1. 数据概览
# =====================================================================
def page_overview():
    st.title("数据概览")

    # 只统计在职员工的打卡数据
    active_df = df_clock[df_clock['name'].isin(active_names)]
    resigned_roster = df_roster[df_roster['status'] == '离职']

    col1, col2, col3 = st.columns(3)
    col1.metric("打卡记录总数", f"{len(active_df):,}")
    col2.metric("在职员工数", f"{len(active_names)} 人")
    col3.metric("数据时间跨度",
                f"{active_df['date'].min().strftime('%Y-%m')} ~ "
                f"{active_df['date'].max().strftime('%Y-%m')}")

    col1, col2 = st.columns(2)
    with col1:
        st.subheader("在职员工部门分布")
        dept_counts = active_roster.groupby('department').size().reset_index(name='人数')
        fig = px.bar(dept_counts, x='department', y='人数',
                     color='人数', color_continuous_scale='Blues')
        fig.update_layout(height=350)
        st.plotly_chart(fig, use_container_width=True)
        st.caption("仅显示在职员工，共 {} 人".format(len(active_roster)))

    with col2:
        st.subheader("打卡状态全局分布")
        morning_dist = active_df['morning_status'].value_counts()
        evening_dist = active_df['evening_status'].value_counts()
        fig = go.Figure()
        fig.add_trace(go.Bar(name='上班', x=morning_dist.index, y=morning_dist.values,
                             marker_color=[MORNING_COLORS.get(k, '#95a5a6') for k in morning_dist.index]))
        fig.add_trace(go.Bar(name='下班', x=evening_dist.index, y=evening_dist.values,
                             marker_color=[EVENING_COLORS.get(k, '#95a5a6') for k in evening_dist.index]))
        fig.update_layout(height=350, barmode='group')
        st.plotly_chart(fig, use_container_width=True)
        st.caption("所有打卡记录的上下班状态汇总。数据来源：打卡时间记录表 clock_records。")

    st.markdown("---")
    st.subheader("考勤状态分类体系")
    st.markdown("""
    | 时段 | 分类 | 判定规则 | 分界说明 |
    |------|------|---------|---------|
    | 上班 | 早到 | 最早打卡时间 < 08:30 | 12:00前打卡 |
    | 上班 | 正常上班 | 08:30 ≤ 最早打卡时间 ≤ 09:00 | 12:00前打卡 |
    | 上班 | 迟到 | 最早打卡时间 > 09:00 | 12:00前打卡 |
    | 上班 | 上班缺卡 | 无12:00前打卡记录 | 仅有下班打卡或无打卡 |
    | 下班 | 早退 | 最晚打卡时间 < 17:30 | 12:00后打卡 |
    | 下班 | 正常下班 | 17:30 ≤ 最晚打卡时间 ≤ 18:00 | 12:00后打卡 |
    | 下班 | 加班 | 最晚打卡时间 > 18:00 | 12:00后打卡 |
    | 下班 | 下班缺卡 | 无12:00后打卡记录 | 仅有上班打卡或无打卡 |
    """)

    # ── 风险预测排行榜 ──
    st.markdown("---")
    st.subheader("离职风险预测排行榜")
    st.caption("基于全体在职员工的集成模型风险概率从高到低排列")

    local_model_page = load_model()
    if local_model_page is not None and len(active_names) >= 3:
        ranking_data = []
        for emp_name in active_names:
            emp_df_pred = df_clock[df_clock['name'] == emp_name].sort_values('date')
            monthly_pred = monthly_agg(emp_df_pred)
            if len(monthly_pred) >= 2:
                result = predict_employee(monthly_pred.to_dict('records'), local_model_page)
                ens_prob = result['ensemble_prob']

                # 计算日均工作时长
                work_hours_list = []
                for _, row in emp_df_pred.iterrows():
                    if row['first_time_str'] and row['last_time_str']:
                        try:
                            t1 = datetime.strptime(row['first_time_str'], '%H:%M')
                            t2 = datetime.strptime(row['last_time_str'], '%H:%M')
                            work_hours_list.append((t2 - t1).total_seconds() / 3600)
                        except (ValueError, TypeError):
                            pass
                avg_hours = np.mean(work_hours_list) if work_hours_list else 0

                emp_roster_row = active_roster[active_roster['name'] == emp_name]
                dept = emp_roster_row['department'].iloc[0] if not emp_roster_row.empty else ''
                ranking_data.append({
                    '姓名': emp_name,
                    '部门': dept,
                    '风险概率': round(ens_prob, 4),
                    '日均工作时长(h)': round(avg_hours, 1),
                })

        if ranking_data:
            df_rank = pd.DataFrame(ranking_data)
            df_rank = df_rank.sort_values('风险概率', ascending=False).reset_index(drop=True)
            df_rank.index = df_rank.index + 1  # 从1开始排名

            # 用颜色标注高风险
            def color_risk(val):
                if val >= 0.6:
                    return 'color: #e74c3c; font-weight: bold'
                elif val >= 0.4:
                    return 'color: #f39c12'
                return 'color: #2ecc71'

            styled = df_rank.style \
                .map(color_risk, subset=['风险概率']) \
                .format({'风险概率': '{:.1%}'})

            col_left, col_right = st.columns([3, 1])
            with col_left:
                st.dataframe(
                    styled,
                    use_container_width=True,
                    height=min(40 * len(df_rank) + 38, 500),
                )
            with col_right:
                st.metric("平均风险", f"{df_rank['风险概率'].mean():.1%}")
                st.metric("最高风险", f"{df_rank['风险概率'].max():.1%}")
                st.caption("风险概率≥60%需重点关注")

            # 风险分布柱状图
            fig_risk = px.histogram(df_rank, x='风险概率', nbins=20,
                                     title='风险概率分布',
                                     color_discrete_sequence=['#3498db'])
            fig_risk.update_layout(height=300, xaxis_tickformat='.0%')
            st.plotly_chart(fig_risk, use_container_width=True)

            # Top 10 高风险人员
            st.subheader("高风险人员（Top 10）")
            top10 = df_rank.head(10).copy()
            top10['排名'] = range(1, len(top10) + 1)
            fig_top = px.bar(top10, x='姓名', y='风险概率', color='部门',
                              text='日均工作时长(h)',
                              title='Top 10 风险人员及日均工作时长',
                              color_discrete_sequence=px.colors.qualitative.Set2)
            fig_top.update_traces(texttemplate='%{text}h', textposition='outside')
            fig_top.update_layout(height=400, yaxis_tickformat='.0%')
            st.plotly_chart(fig_top, use_container_width=True)
            st.caption("柱高=离职风险概率，标签=日平均工作时长")
    else:
        if local_model_page is None:
            st.info("模型未训练，上传足够数据后自动训练以启用风险预测。")
        else:
            st.info("在职员工不足3人，暂不显示排行榜。")


# =====================================================================
#  2. 部门考勤看板（可筛选年份+月份）
# =====================================================================
def page_dashboard():
    st.title("部门考勤看板")

    active_df = df_clock[df_clock['name'].isin(active_names)]

    years = sorted(active_df['year'].unique(), reverse=True)
    months = sorted(active_df['month'].unique())
    depts = sorted(active_roster['department'].unique())

    col1, col2, col3 = st.columns(3)
    with col1:
        sel_year = st.selectbox("选择年份", ['全部'] + [str(y) for y in years], index=0)
    with col2:
        sel_month = st.selectbox("选择月份", ['全部'] + [str(m) + '月' for m in months], index=0)
    with col3:
        sel_dept = st.selectbox("选择部门", ['全部'] + depts, index=0)

    # 筛选
    filtered = active_df.copy()
    if sel_year != '全部':
        filtered = filtered[filtered['year'] == int(sel_year)]
    if sel_month != '全部':
        filtered = filtered[filtered['month'] == int(sel_month.replace('月', ''))]
    if sel_dept != '全部':
        emp_in_dept = active_roster[active_roster['department'] == sel_dept]['name'].unique()
        filtered = filtered[filtered['name'].isin(emp_in_dept)]

    st.caption(
        f"筛选条件：年份={sel_year}，月份={sel_month}，部门={sel_dept}。"
        f"当前筛选后 {len(filtered)} 条记录。"
    )

    col1, col2 = st.columns(2)
    with col1:
        st.subheader("上班打卡状态分布")
        m_dist = filtered['morning_status'].value_counts().reset_index()
        m_dist.columns = ['状态', '次数']
        fig = px.bar(m_dist, x='状态', y='次数',
                     color='状态',
                     color_discrete_map=MORNING_COLORS)
        fig.update_layout(height=350, showlegend=False)
        st.plotly_chart(fig, use_container_width=True)

    with col2:
        st.subheader("下班打卡状态分布")
        e_dist = filtered['evening_status'].value_counts().reset_index()
        e_dist.columns = ['状态', '次数']
        fig = px.bar(e_dist, x='状态', y='次数',
                     color='状态',
                     color_discrete_map=EVENING_COLORS)
        fig.update_layout(height=350, showlegend=False)
        st.plotly_chart(fig, use_container_width=True)

    # 年度趋势
    st.subheader("月度打卡趋势")
    agg_monthly = filtered.groupby(['year', 'month']).agg(
        总记录数=('name', 'count'),
        早到率=('morning_status', lambda x: (x == '早到').mean()),
        迟到率=('morning_status', lambda x: (x == '迟到').mean()),
        加班率=('evening_status', lambda x: (x == '加班').mean()),
        早退率=('evening_status', lambda x: (x == '早退').mean()),
    ).reset_index()
    agg_monthly['时间'] = agg_monthly.apply(lambda r: f"{int(r['year'])}-{int(r['month']):02d}", axis=1)

    if len(agg_monthly) > 0:
        fig = go.Figure()
        for metric, color, name in [('早到率', '#3498db', '早到'), ('迟到率', '#e74c3c', '迟到'),
                                     ('加班率', '#9b59b6', '加班'), ('早退率', '#f39c12', '早退')]:
            fig.add_trace(go.Scatter(x=agg_monthly['时间'], y=agg_monthly[metric],
                                     mode='lines+markers', name=name, line=dict(color=color)))
        fig.update_layout(height=400, xaxis_title='月份', yaxis_title='比例',
                          yaxis=dict(tickformat='.0%'))
        st.plotly_chart(fig, use_container_width=True)
        st.caption("趋势线基于筛选后数据计算，比例为当月该状态次数 / 当月总记录数。")

    # 部门横向对比
    if sel_dept == '全部':
        st.subheader("各部门考勤指标对比")
        dept_compare = []
        for d in depts:
            emp_ids = active_roster[active_roster['department'] == d]['name'].unique()
            d_df = active_df[active_df['name'].isin(emp_ids)]
            if len(d_df) == 0:
                continue
            dept_compare.append({
                '部门': d,
                '记录数': len(d_df),
                '早到率': (d_df['morning_status'] == '早到').mean(),
                '迟到率': (d_df['morning_status'] == '迟到').mean(),
                '加班率': (d_df['evening_status'] == '加班').mean(),
                '早退率': (d_df['evening_status'] == '早退').mean(),
                '缺卡率': (d_df['evening_status'] == '下班缺卡').mean(),
            })
        if dept_compare:
            df_dc = pd.DataFrame(dept_compare)
            fig = px.bar(df_dc, x='部门', y=['早到率', '迟到率', '加班率', '早退率', '缺卡率'],
                         barmode='group', title='各部门考勤指标对比')
            fig.update_layout(height=400, yaxis_tickformat='.0%')
            st.plotly_chart(fig, use_container_width=True)
            st.caption("基于所有在职员工打卡记录统计。迟到/早退率越低、加班率越低的部门整体考勤越好。")


# =====================================================================
#  3. 员工考勤分析（分部门，仅在职）
# =====================================================================
def page_employee():
    st.title("员工考勤分析")
    st.caption("仅显示在职员工。可搜索全公司员工，或按部门筛选查看。")

    active_df = df_clock[df_clock['name'].isin(active_names)]

    # 全公司员工搜索（放在最前面）
    all_emp_names = sorted(active_names)
    if 'emp_search' not in st.session_state:
        st.session_state.emp_search = ''
    emp_search = st.text_input("🔍 搜索员工姓名（全公司）", value=st.session_state.emp_search,
                                placeholder="输入姓名快速定位员工...", key="emp_search_input")
    st.session_state.emp_search = emp_search
    searched_emps = [e for e in all_emp_names if emp_search.lower() in e.lower()]
    if not searched_emps:
        searched_emps = all_emp_names

    depts = sorted(active_roster['department'].unique())

    # 两个页面共享同一个 session_state
    if 'shared_dept' not in st.session_state:
        st.session_state.shared_dept = depts[0]
    if 'shared_emp' not in st.session_state:
        st.session_state.shared_emp = searched_emps[0] if searched_emps else None

    # 如果当前选中的员工不在搜索结果中，自动跳到第一个
    if st.session_state.shared_emp not in searched_emps:
        st.session_state.shared_emp = searched_emps[0] if searched_emps else None

    # 根据选中员工自动确定部门
    if st.session_state.shared_emp:
        emp_dept = active_roster[active_roster['name'] == st.session_state.shared_emp]['department'].iloc[0]
        if emp_dept in depts:
            st.session_state.shared_dept = emp_dept

    col1, col2 = st.columns(2)
    with col1:
        sel_dept = st.selectbox("选择部门", depts, key="shared_dept")
    with col2:
        # 部门筛选后只显示该部门且在搜索结果中的员工
        dept_emps = active_roster[active_roster['department'] == sel_dept]['name'].unique()
        dept_filtered = sorted([e for e in dept_emps if e in searched_emps])
        if not dept_filtered:
            dept_filtered = sorted(dept_emps)
        if st.session_state.shared_emp not in dept_filtered:
            st.session_state.shared_emp = dept_filtered[0] if dept_filtered else None
        sel_emp = st.selectbox("选择员工", dept_filtered, key="shared_emp")
    emp_df = active_df[active_df['name'] == sel_emp].sort_values('date')
    emp_roster = active_roster[active_roster['name'] == sel_emp].iloc[0]

    st.markdown(f"**{sel_emp}** | {emp_roster['department']} | {emp_roster['rank']}")
    st.caption(f"数据范围：{emp_df['date'].min().strftime('%Y-%m-%d')} ~ {emp_df['date'].max().strftime('%Y-%m-%d')}，共 {len(emp_df)} 条打卡记录")

    # 本月考勤柱状图
    st.subheader("本月考勤分布")
    latest_date = emp_df['date'].max()
    cur_year, cur_month = latest_date.year, latest_date.month
    cur_df = emp_df[(emp_df['year'] == cur_year) & (emp_df['month'] == cur_month)]
    if len(cur_df) > 0:
        col1, col2 = st.columns(2)
        with col1:
            m_fig = px.bar(
                x=cur_df['morning_status'].value_counts().index,
                y=cur_df['morning_status'].value_counts().values,
                title=f"上班 ({cur_year}-{cur_month:02d})",
                color=cur_df['morning_status'].value_counts().index,
                color_discrete_map=MORNING_COLORS,
                labels={'x': '状态', 'y': '天数'},
            )
            m_fig.update_layout(showlegend=False, height=300)
            st.plotly_chart(m_fig, use_container_width=True)
        with col2:
            e_fig = px.bar(
                x=cur_df['evening_status'].value_counts().index,
                y=cur_df['evening_status'].value_counts().values,
                title=f"下班 ({cur_year}-{cur_month:02d})",
                color=cur_df['evening_status'].value_counts().index,
                color_discrete_map=EVENING_COLORS,
                labels={'x': '状态', 'y': '天数'},
            )
            e_fig.update_layout(showlegend=False, height=300)
            st.plotly_chart(e_fig, use_container_width=True)
        st.caption(f"当前月份：{cur_year}年{cur_month}月。柱状图展示该月各考勤状态的天数分布。")
    else:
        st.info("本月暂无打卡数据。")

    # 年度考勤趋势
    st.subheader("年度打卡趋势")
    emp_monthly = monthly_agg(emp_df)
    if len(emp_monthly) > 0:
        emp_monthly['时间'] = emp_monthly.apply(lambda r: f"{int(r['year'])}-{int(r['month']):02d}", axis=1)
        fig = go.Figure()
        for metric, color, name in [('early_arr_rate', '#3498db', '早到率'),
                                     ('late_arr_rate', '#e74c3c', '迟到率'),
                                     ('overtime_rate', '#9b59b6', '加班率'),
                                     ('early_leave_rate', '#f39c12', '早退率')]:
            fig.add_trace(go.Scatter(
                x=emp_monthly['时间'], y=emp_monthly[metric],
                mode='lines+markers', name=name, line=dict(color=color, width=2),
            ))
        fig.update_layout(height=400, xaxis_title='月份', yaxis_title='比例',
                          yaxis=dict(tickformat='.0%'))
        st.plotly_chart(fig, use_container_width=True)
        st.caption("月度趋势线展示该员工各月早到/迟到/加班/早退的比例变化。环比变化可反映行为趋势。")

    # 全部打卡记录明细
    st.subheader("全部打卡记录")
    display = emp_df[['date', 'first_time_str', 'last_time_str',
                       'morning_status', 'evening_status']].copy()
    display.columns = ['日期', '上班打卡', '下班打卡', '上班状态', '下班状态']
    display['日期'] = display['日期'].dt.strftime('%Y-%m-%d')
    st.dataframe(display, use_container_width=True, hide_index=True)
    st.caption(f"共 {len(display)} 条记录。仅展示当前员工的在职打卡时间数据。")


# =====================================================================
#  4. 离职风险预测（相对百分位排名）
# =====================================================================
def page_prediction():
    st.title("离职风险预测")
    st.caption(
        "基于在职员工群体内部的相对风险百分位排名。"
        "三模型集成（RF + XGBoost + Logistic Regression），"
        "结合离职前行为变化、同部门偏差、个人趋势三大视角。"
    )

    active_df = df_clock[df_clock['name'].isin(active_names)]

    # 每次进入预测页面都重新加载模型（确保上传新数据后模型已更新）
    local_model = load_model()
    if local_model is None:
        st.error("模型文件未找到，请先运行 train_models.py 训练模型。")
        return

    depts = sorted(active_roster['department'].unique())

    # 全公司员工搜索（放在最前面）
    all_emp_names = sorted(active_names)
    if 'pred_emp_search' not in st.session_state:
        st.session_state.pred_emp_search = ''
    pred_emp_search = st.text_input("🔍 搜索员工姓名（全公司）", value=st.session_state.pred_emp_search,
                                    placeholder="输入姓名快速定位员工...", key="pred_emp_search_input")
    st.session_state.pred_emp_search = pred_emp_search
    searched_emps = [e for e in all_emp_names if pred_emp_search.lower() in e.lower()]
    if not searched_emps:
        searched_emps = all_emp_names

    # 与员工考勤分析页面共享同一组 session_state
    if 'shared_dept' not in st.session_state:
        st.session_state.shared_dept = depts[0]
    if 'shared_emp' not in st.session_state:
        st.session_state.shared_emp = searched_emps[0] if searched_emps else None

    # 如果当前选中的员工不在搜索结果中，自动跳到第一个
    if st.session_state.shared_emp not in searched_emps:
        st.session_state.shared_emp = searched_emps[0] if searched_emps else None

    # 根据选中员工自动确定部门
    if st.session_state.shared_emp:
        emp_dept = active_roster[active_roster['name'] == st.session_state.shared_emp]['department'].iloc[0]
        if emp_dept in depts:
            st.session_state.shared_dept = emp_dept

    col1, col2 = st.columns(2)
    with col1:
        sel_dept = st.selectbox("选择部门", depts, key="shared_dept")
    with col2:
        # 部门筛选后只显示该部门且在搜索结果中的员工
        dept_emps = active_roster[active_roster['department'] == sel_dept]['name'].unique()
        dept_filtered = sorted([e for e in dept_emps if e in searched_emps])
        if not dept_filtered:
            dept_filtered = sorted(dept_emps)
        if st.session_state.shared_emp not in dept_filtered:
            st.session_state.shared_emp = dept_filtered[0] if dept_filtered else None
        sel_emp = st.selectbox("选择员工", dept_filtered, key="shared_emp")

    emp_df = active_df[active_df['name'] == sel_emp].sort_values('date')
    monthly = monthly_agg(emp_df)
    if len(monthly) < 2:
        st.warning("该员工数据不足（至少需要2个月数据）。")
        return

    # 单员工预测
    monthly_dict = monthly.to_dict('records')
    result = predict_employee(monthly_dict, local_model)
    ens_prob = result['ensemble_prob']

    # 计算在职员工群体的基准分布
    eligible = [n for n in active_names
                if len(active_df[active_df['name'] == n]) >= 20]
    base_probs = []
    for n in eligible:
        ed = active_df[active_df['name'] == n].sort_values('date')
        md = monthly_agg(ed)
        if len(md) >= 2:
            r = predict_employee(md.to_dict('records'), local_model)
            base_probs.append(r['ensemble_prob'])

    if base_probs:
        all_probs = sorted(base_probs)
        rank = sum(1 for p in all_probs if p <= ens_prob)
        percentile = rank / len(all_probs) * 100
    else:
        percentile = 50.0

    # 显示
    col1, col2, col3, col4 = st.columns(4)
    models = [('RF', result['rf_prob'], '#3498db'),
              ('XGBoost', result['xgb_prob'], '#e74c3c'),
              ('LR', result['lr_prob'], '#2ecc71'),
              ('集成', ens_prob, '#9b59b6')]

    for col, (name, prob, color) in zip([col1, col2, col3, col4], models):
        with col:
            rank_label = ""
            if name == '集成':
                if percentile >= 85:
                    rank_label = "高风险"
                    bg_color = "#e74c3c"
                elif percentile >= 60:
                    rank_label = "中高风险"
                    bg_color = "#f39c12"
                elif percentile >= 30:
                    rank_label = "中等风险"
                    bg_color = "#3498db"
                else:
                    rank_label = "低风险"
                    bg_color = "#2ecc71"
                col.markdown(f"""
                <div style="text-align:center;padding:20px;border-radius:10px;border:2px solid {color};background:rgba(0,0,0,0.02);">
                    <h4 style="color:{color};">{name}</h4>
                    <h1 style="color:{color};">{prob:.1%}</h1>
                    <p style="font-size:16px;color:{bg_color};font-weight:bold;">{rank_label}</p>
                    <p style="font-size:13px;color:#666;">排名: P{percentile:.0f}</p>
                </div>
                """, unsafe_allow_html=True)
            else:
                col.markdown(f"""
                <div style="text-align:center;padding:20px;border-radius:10px;border:2px solid {color};background:rgba(0,0,0,0.02);">
                    <h4 style="color:{color};">{name}</h4>
                    <h1 style="color:{color};">{prob:.1%}</h1>
                </div>
                """, unsafe_allow_html=True)

    # 相对排名说明
    st.markdown("---")

    if base_probs:
        st.subheader("相对风险排名")
        st.caption(f"基于 {len(base_probs)} 名在职员工的集成风险概率分布。")
        fig = go.Figure()
        fig.add_trace(go.Histogram(x=all_probs, nbinsx=20, name='在职员工分布',
                                   marker_color='rgba(52, 152, 219, 0.6)'))
        fig.add_vline(x=ens_prob, line_dash='dash', line_color='#e74c3c',
                      annotation_text=f"{sel_emp} ({ens_prob:.1%})")
        fig.update_layout(height=350, xaxis_title='离职风险概率', yaxis_title='人数')
        st.plotly_chart(fig, use_container_width=True)

    # 风险等级解读
    st.subheader("风险等级说明")
    st.markdown("""
    | 等级 | 百分位范围 | 含义 |
    |------|----------|------|
    | 🔴 高风险 | ≥ 85% | 打卡行为偏离同部门平均水平明显，需关注 |
    | 🟠 中高风险 | 60% ~ 84% | 有部分预警信号，建议日常沟通 |
    | 🔵 中等风险 | 30% ~ 59% | 与部门平均水准接近，持续观察 |
    | 🟢 低风险 | < 30% | 打卡行为稳定，风险较低 |
    """)

    # 特征解释
    st.subheader("模型使用的考勤特征")
    st.markdown("""
    **三大分析视角：**
    1. **离职前对比**：近3个月 vs 此前稳定期的打卡行为偏差（迟到率变化、加班率变化等）
    2. **同部门对比**：个人行为与部门同期平均水平的偏离度
    3. **个人趋势**：打卡行为波动性（标准差）、迟到/早退/加班频率变化趋势
    """)


# =====================================================================
#  5. 模型信息
# =====================================================================
def page_model_info():
    st.title("模型信息")

    local_model = load_model()
    if local_model is None:
        st.error("模型文件未找到，请先运行 train_models.py。")
        return

    # 显示模型文件最后修改时间（方便确认模型已更新）
    mtime_str = "未知"
    if os.path.exists(MODEL_PATH):
        mtime = os.path.getmtime(MODEL_PATH)
        mtime_str = datetime.fromtimestamp(mtime).strftime('%Y-%m-%d %H:%M:%S')

    st.markdown(f"**训练日期**: {local_model.get('training_date', '未知')}")
    st.markdown(f"**模型文件更新于**: {mtime_str}")
    st.markdown(f"**训练样本数**: {local_model.get('n_samples', '?')} 名员工")
    st.markdown(f"**特征维度**: {local_model.get('n_features', '?')} 个")

    st.subheader("模型简介")
    st.markdown("""
    | 模型 | 类型 | 特点 | 适用场景 |
    |------|------|------|---------|
    | **Random Forest** | 集成学习（Bagging） | 多棵决策树投票，抗过拟合强，可输出特征重要性 | 小样本、非线性关系、需要可解释性 |
    | **XGBoost** | 梯度提升（Boosting） | 迭代优化，支持样本权重，L1+L2正则化 | 处理类别不平衡，捕捉复杂交互效应 |
    | **Logistic Regression** | 线性模型 | 简单可解释，L2正则化（C=0.1） | 基线对照、线性关系、低方差需求 |
    """)

    st.subheader("评估指标说明")
    st.markdown("""
    | 指标 | 含义 | 范围 | 说明 |
    |------|------|:----:|------|
    | **CV AUC (3折)** | 3折交叉验证的 ROC-AUC 均值 | 0.5~1.0 | 最可靠的评估指标，>0.7 即有区分度 |
    | **测试 F1** | 测试集上精确率与召回率的调和平均 | 0~1 | 越高表示分类越准确 |
    | **测试 AUC** | 测试集上的 ROC 曲线下面积 | 0.5~1.0 | >0.8 为良好，>0.9 为优秀 |
    """)
    st.caption(f"小样本场景（{local_model.get('n_samples', '?')}人）下，CV AUC 比测试集分数更有参考价值。")

    st.subheader("模型性能")
    metrics = []
    for model_name in ['rf_metrics', 'xgb_metrics', 'lr_metrics']:
        m = local_model.get(model_name, {})
        label = {'rf_metrics': 'Random Forest', 'xgb_metrics': 'XGBoost', 'lr_metrics': 'LR'}[model_name]
        cv_auc = m.get('cv_auc')
        cv_str = f"{cv_auc:.3f}" if cv_auc is not None and not (isinstance(cv_auc, float) and np.isnan(cv_auc)) else "—"
        metrics.append({
            '模型': label,
            'CV AUC (3折)': cv_str,
            '测试 F1': f"{m.get('f1', 0):.3f}",
            '测试 AUC': f"{m.get('auc', 0):.3f}",
        })
    df_m = pd.DataFrame(metrics)
    st.dataframe(df_m, use_container_width=True, hide_index=True)
    n_test = local_model.get('n_test', '?')
    st.caption(f"CV AUC = 3折交叉验证，比测试集分数更可靠（测试集仅{n_test}个样本，波动大）。")

    # ── 10 次重复划分评估汇总 ──
    rep = local_model.get('repeat_metrics', None)
    if rep:
        st.subheader("10 次重复随机划分测试集评估")
        st.caption("每次完整流水线：划分→选特征→缩放→训练→测试，10次不同 seed 的均值 ± 标准差。")
        rep_rows = []
        for label, key in [('Random Forest', 'rf'), ('XGBoost', 'xgb'), ('Logistic Regression', 'lr')]:
            aucs = rep[key]['test_auc']
            f1s = rep[key]['test_f1']
            rep_rows.append({
                '模型': label,
                '测试 AUC (均值±标准差)': f"{np.mean(aucs):.3f} ± {np.std(aucs):.3f}",
                '测试 F1 (均值±标准差)': f"{np.mean(f1s):.3f} ± {np.std(f1s):.3f}",
            })
        st.dataframe(pd.DataFrame(rep_rows), use_container_width=True, hide_index=True)

    st.subheader("模型设计")
    # 动态计算实际数据
    n_act = local_model.get('n_active', '?')
    n_res = local_model.get('n_resigned', '?')
    rf_cv = local_model.get('rf_metrics', {}).get('cv_auc', None)
    xgb_cv = local_model.get('xgb_metrics', {}).get('cv_auc', None)
    lr_cv = local_model.get('lr_metrics', {}).get('cv_auc', None)
    cv_vals = [v for v in [rf_cv, xgb_cv, lr_cv] if v is not None and not (isinstance(v, float) and np.isnan(v))]
    cv_range = f"{min(cv_vals):.2f}~{max(cv_vals):.2f}" if len(cv_vals) >= 2 else "?"
    # 权重
    w = local_model.get('ensemble_weights', None)
    if w and not any(np.isnan(v) for v in [w['rf'], w['xgb'], w['lr']]):
        total = w['rf'] + w['xgb'] + w['lr']
        w_rf, w_xgb, w_lr = w['rf']/total, w['xgb']/total, w['lr']/total
        w_str = f"RF:{w['rf']:.3f}, XGB:{w['xgb']:.3f}, LR:{w['lr']:.3f} → 归一化权重 {w_rf:.3f}/{w_xgb:.3f}/{w_lr:.3f}"
    else:
        w_str = "等权平均"

    st.markdown(f"""
    | 设计要素 | 具体措施 | 目的 |
    |------|---------|------|
    | **特征体系** | 22个核心特征（RF预选Top22，从35维压缩37%），三大视角：离职前vs稳定期偏差、同部门偏离度、个人趋势波动 + 行为协同模式 + 恶化斜率 | 特征缩减为主要正则化手段，保留信号多样性 |
    | **过拟合防护** | 特征砍掉37%（35→22）+ 三模型集成 + 3折CV + 样本权重 | 特征缩减 + 模型集成双重防过拟合，CV AUC {cv_range} |
    | **类别平衡** | RF(class_weight='balanced'), XGB(sample_weight), LR(class_weight='balanced') | 解决在职/离职样本不均衡（{n_act}:{n_res}） |
    | **模型集成** | 三模型按 CV AUC 加权平均（{w_str}） | 高精度模型获得更大权重，提升集成效果 |
    | **评估策略** | 3折交叉验证（主要指标） + 测试集验证 | CV AUC 比测试集分数更可靠（测试集仅{n_test}个样本） |
    | **风险排名** | 基于全体在职员工的百分位相对排名 | 消除绝对概率的不可比性，聚焦相对风险 |
    """)

    st.subheader("特征重要性（RF Top 10）")
    # 提取特征重要性
    rf_model = local_model.get('rf_model')
    if rf_model and hasattr(rf_model, 'feature_importances_'):
        fc = local_model['feature_cols']
        # 特征中文映射
        FEATURE_CN = {
            'total_months': '总在职月数',
            'recent_early_arr': '近期早到率',
            'recent_late_arr': '近期迟到率',
            'recent_early_leave': '近期早退率',
            'recent_overtime': '近期加班率',
            'recent_miss_eve': '近期下班缺卡率',
            'recent_avg_morning': '近期平均上班时间',
            'recent_avg_evening': '近期平均下班时间',
            'last_month_late_arr': '末月迟到率',
            'last_month_overtime': '末月加班率',
            'last_month_miss_eve': '末月缺卡率',
            'trend_early_arr': '早到率变化趋势',
            'trend_late_arr': '迟到率变化趋势',
            'trend_early_leave': '早退率变化趋势',
            'trend_overtime': '加班率变化趋势',
            'trend_miss_eve': '缺卡率变化趋势',
            'slope_late_arr': '迟到率恶化斜率',
            'slope_overtime': '加班率恶化斜率',
            'slope_miss_eve': '缺卡率恶化斜率',
            'std_early_arr': '早到率波动性',
            'std_late_arr': '迟到率波动性',
            'std_early_leave': '早退率波动性',
            'std_overtime': '加班率波动性',
            'cv_late_arr': '迟到率变异系数',
            'cv_overtime': '加班率变异系数',
            'pos_pattern_rate': '积极行为率（早到+加班）',
            'neg_pattern_rate': '消极行为率（迟到+早退）',
            'damage_control_rate': '补偿行为率（迟到+加班）',
            'full_absent_rate': '全天缺卡率',
            'dev_early_arr': '与部门早到率偏差',
            'dev_late_arr': '与部门迟到率偏差',
            'dev_early_leave': '与部门早退率偏差',
            'dev_overtime': '与部门加班率偏差',
            'dev_miss': '与部门缺卡率偏差',
            'department_encoded': '部门编码',
        }
        imp = pd.DataFrame({'特征': fc, '重要性': rf_model.feature_importances_})
        imp['特征'] = imp['特征'].map(FEATURE_CN).fillna(imp['特征'])
        imp = imp.sort_values('重要性', ascending=False).head(10)
        fig = px.bar(imp, x='重要性', y='特征', orientation='h', color='重要性',
                     color_continuous_scale='Blues')
        fig.update_layout(height=350, yaxis={'categoryorder': 'total ascending'})
        st.plotly_chart(fig, use_container_width=True)
        st.caption("特征重要性反映各考勤指标对离职预测的贡献度。越长的条表示该特征越重要。")


# =====================================================================
#  路由
# =====================================================================
if active_names:
    if page == "数据概览":
        page_overview()
    elif page == "部门考勤看板":
        page_dashboard()
    elif page == "员工考勤分析":
        page_employee()
    elif page == "离职风险预测":
        page_prediction()
    elif page == "模型信息":
        page_model_info()
else:
    st.title("👋 欢迎使用考勤分析系统")
    st.markdown("""
    ### 开始之前

    这是首次运行，数据库中没有数据。请按以下步骤操作：

    1. **下载模板** ← 在左侧「格式说明 & 下载模板」中下载
    2. **准备数据**：按模板格式填写考勤数据
    3. **上传数据**：使用左侧的「上传」功能上传 Excel 文件

    上传完成后，页面将自动刷新并显示分析看板。
    """)
